'''
	Script to generate personalized student exams, and a blank one.

	Execution Requirements:
	- Python 3.5+
	- openpyxl (et-xmlfile, jdcal): all three are installed with 'sudo pip install openpyxl'

	Assumptions:
	- The script can re-generate student.tex and compile exam.tex to get a PDF.
	- The file zulassingsliste.xlsx exists and has the following format:
			All data on sheet 1
			Columns with "Matrikelnummer", "Vorname" and "Name" exist (not case sensitive)
			No empty rows (or pdf-s with None will be generated)
			...
'''
import os, shutil, os.path
# from openpyxl import Workbook
from openpyxl import load_workbook

# Target folder for exam PDF files.
folder="pdf"

def genexam(lastname, firstname, studentid):
	'''
		Generates an exam for the given student, if not existing.
		Target file is stored in "folder", with the student ID as file name.
	'''
	targetfile = folder+os.sep+studentid+'.pdf'
	if os.path.isfile(targetfile):
		print("Skipping generation of %s, because it exists ..."%targetfile)
		return
	else:
		print("Generating "+targetfile)
	studentfile=open('student.tex','w')
	studentfile.write('''
\\newcommand{\\studentid}{%s}\n
\\newcommand{\\studentname}{%s %s}
	'''%(studentid, firstname, lastname))
	studentfile.close()
	os.system('latexmk -C')
	os.system('lualatex exam.tex')  # double run needed by exam class
	os.system('lualatex exam.tex')
	shutil.move('exam.pdf', targetfile)

try:
	os.mkdir(folder)
except:
	pass

# older code for CSV file format
# with open('teilnehmer.csv', 'rb') as csvfile:
# 	reader = csv.reader(csvfile, delimiter=';')
# 	for row in reader:
# 		genexam(*row)
#

# load the xlsx file for reading only and extract data from the first sheet
wb = load_workbook(filename='zulassungsliste.xlsx', read_only=True)
sheets = wb.get_sheet_names()
ws = wb[sheets[0]]

startRow = 1
matrikelCol = 1
nameCol = 1
vornameCol = 1

matrikelFound = False
nameFound = False
vornameFound = False

# find the starting row for the data and the columns for id, name and first name
for i in range(1,ws.max_row+1):
	for j in range(1, ws.max_column+1):
		if 'matrikelnummer' == str(ws.cell(row=i,column=j).value).lower():
			startRow = i+1
			matrikelCol = j
			matrikelFound = True
		elif 'name' == str(ws.cell(row=i, column=j).value).lower():
			nameCol = j
			nameFound = True
		elif 'vorname' == str(ws.cell(row=i, column=j).value).lower():
			vornameCol = j
			vornameFound = True
		# print(ws.cell(row=i,column=j).value)

# if all necessary columns found, generate exams for every student
if matrikelFound and nameFound and vornameFound:
	# generate students as a list of tuples and generate corresponding exam sheets
	# studentsList = []

	for i in range(startRow,ws.max_row+1):
		student = (str(ws.cell(row=i,column=matrikelCol).value), ws.cell(row=i,column=vornameCol).value, ws.cell(row=i,column=nameCol).value)
		print(student)
		# studentsList.append(student)
		genexam(studentid=str(ws.cell(row=i, column=matrikelCol).value), firstname=ws.cell(row=i, column=vornameCol).value, lastname=ws.cell(row=i, column=nameCol).value)

	# Also generate a blank one. LaTex needs some content, so we use gibberish white space.
	genexam('~','~','~')
else:
	print('At least one of the needed columns has not been found! Check your .xlsx input file format!')